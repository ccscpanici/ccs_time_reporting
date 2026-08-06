"""Small dependency-free factories used by the Django test suite.

These helpers intentionally avoid factory-boy/model-bakery so the core suite can
run with only the production requirements installed.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from decimal import Decimal
from typing import Any

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group

from accounts.models import EmployeeProfile
from timesheets.models import Customer, Job, TimeEntry, Timesheet
from openpyxl import Workbook

User = get_user_model()


def make_user(*, username: str, password: str = "test-password", **overrides: Any):
    values = {
        "first_name": username.replace("_", " ").title(),
        "last_name": "User",
        "email": f"{username}@gotoccs.com",
    }
    values.update(overrides)
    return User.objects.create_user(username=username, password=password, **values)


def get_group(name: str) -> Group:
    group, _ = Group.objects.get_or_create(name=name)
    return group


def add_user_to_group(user, name: str) -> Group:
    group = get_group(name)
    user.groups.add(group)
    return group


def make_employee_profile(*, user, supervisor=None, **overrides: Any) -> EmployeeProfile:
    values = {"supervisor": supervisor}
    values.update(overrides)
    return EmployeeProfile.objects.create(user=user, **values)


def make_customer(*, name: str = "Test Customer", **overrides: Any) -> Customer:
    return Customer.objects.create(name=name, **overrides)


def make_job(*, job_number: str = "26001", **overrides: Any) -> Job:
    values = {
        "description": f"Description for {job_number}",
        "year": 2026,
        "job_month": 1,
        "job_status": Job.STATUS_ACTIVE,
        "invoice_status": Job.INVOICE_STATUS_PROGRESS,
        "active": True,
    }
    values.update(overrides)
    return Job.objects.create(job_number=job_number, **values)


def make_timesheet(
    *,
    employee,
    week_start: date,
    status: str = Timesheet.Status.DRAFT,
    **overrides: Any,
) -> Timesheet:
    values = {
        "mileage_rate": Decimal("0.72"),
        "entries_per_day": 5,
        "template_entries_per_day": 5,
    }
    values.update(overrides)
    return Timesheet.objects.create(
        employee=employee,
        week_start=week_start,
        status=status,
        **values,
    )


def make_time_entry(
    *,
    timesheet: Timesheet,
    work_date: date | None = None,
    row_order: int = 1,
    **overrides: Any,
) -> TimeEntry:
    values = {
        "job_number": "26001",
        "regular_hours": Decimal("0.00"),
        "overtime_hours": Decimal("0.00"),
        "doubletime_hours": Decimal("0.00"),
        "description": "Test work",
    }
    values.update(overrides)
    return TimeEntry.objects.create(
        timesheet=timesheet,
        work_date=work_date or timesheet.week_start,
        row_order=row_order,
        **values,
    )


def write_job_workbook(
    path: str | Path,
    *,
    headers: list[str],
    rows: list[list[Any]],
    sheet_title: str = "Jobs - Quotes",
    preface_rows: list[list[Any]] | None = None,
) -> Path:
    """Write a compact XLSX workbook for job-import tests and return its path."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title

    for row in preface_rows or []:
        worksheet.append(row)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    output_path = Path(path)
    workbook.save(output_path)
    return output_path


def write_timesheet_workbook(
    path: str | Path,
    *,
    week_start: date | None = date(2026, 8, 2),
    time_rows: list[dict[str, Any]] | None = None,
    expense_rows: list[dict[str, Any]] | None = None,
    part_rows: list[dict[str, Any]] | None = None,
    include_expense_sheet: bool = True,
    include_parts_sheet: bool = True,
) -> Path:
    """Write a compact workbook using the production timesheet cell mapping."""
    workbook = Workbook()
    time_sheet = workbook.active
    time_sheet.title = "Time Sheet"

    if week_start is not None:
        time_sheet["F7"] = week_start

    for item in time_rows or []:
        row = int(item.get("row", 20))
        if "date" in item:
            time_sheet[f"A{row}"] = item["date"]
        time_sheet[f"B{row}"] = item.get("job_number", "")
        time_sheet[f"C{row}"] = item.get("work_code", "")
        time_sheet[f"D{row}"] = item.get("regular", 0)
        time_sheet[f"E{row}"] = item.get("overtime", 0)
        time_sheet[f"F{row}"] = item.get("doubletime", 0)
        time_sheet[f"G{row}"] = item.get("description", "")
        if "overnight" in item:
            chunk_end = 24 + ((row - 20) // 5) * 5
            time_sheet[f"N{chunk_end}"] = item["overnight"]

    if include_expense_sheet:
        expense_sheet = workbook.create_sheet("Expense Report")
        for item in expense_rows or []:
            row = int(item.get("row", 9))
            expense_sheet[f"C{row}"] = item.get("miles", 0)
            expense_sheet[f"E{row}"] = item.get("per_diem_food", 0)
            expense_sheet[f"F{row}"] = item.get("air_fare", 0)
            expense_sheet[f"G{row}"] = item.get("hotel", 0)
            expense_sheet[f"H{row}"] = item.get("tolls_parking", 0)
            expense_sheet[f"I{row}"] = item.get("rental_car_fuel", 0)
            expense_sheet[f"J{row}"] = item.get("business_meals", 0)
            expense_sheet[f"K{row}"] = item.get("other_expense", 0)
            expense_sheet[f"L{row}"] = item.get("explanation", "")

    if include_parts_sheet:
        parts_sheet = workbook.create_sheet("Parts Report")
        for item in part_rows or []:
            row = int(item.get("row", 9))
            parts_sheet[f"B{row}"] = item.get("ee_stock_job_number", "")
            parts_sheet[f"C{row}"] = item.get("quantity", 0)
            parts_sheet[f"D{row}"] = item.get("description", "")
            parts_sheet[f"E{row}"] = item.get("notes", "")
            parts_sheet[f"F{row}"] = item.get("reorder", False)

    output_path = Path(path)
    workbook.save(output_path)
    return output_path
