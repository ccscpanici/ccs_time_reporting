from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from django.db import transaction
from openpyxl import load_workbook
from ..models import Expense, Job, MileageRate, PartEntry, TimeEntry, Timesheet, WorkCode
from .helpers import as_decimal, parse_bool_cell
from .workbook_mapping import *


@dataclass
class ParsedTimeEntry:
    work_date: date
    row_order: int
    job_number: str
    work_code: str
    regular_hours: Decimal = Decimal("0")
    overtime_hours: Decimal = Decimal("0")
    doubletime_hours: Decimal = Decimal("0")
    overnight_stay: bool = False
    description: str = ""


@dataclass
class ParsedExpenseEntry:
    time_sheet_row: int
    work_date: date
    row_order: int
    miles: Decimal = Decimal("0")
    per_diem_food: Decimal = Decimal("0")
    air_fare: Decimal = Decimal("0")
    hotel: Decimal = Decimal("0")
    tolls_parking: Decimal = Decimal("0")
    rental_car_fuel: Decimal = Decimal("0")
    business_meals: Decimal = Decimal("0")
    other_expense: Decimal = Decimal("0")
    explanation_of_expenses: str = ""


@dataclass
class ParsedPartEntry:
    time_sheet_row: int
    work_date: date
    row_order: int
    ee_stock_job_number: str = ""
    quantity: Decimal = Decimal("0")
    part_description_part_number: str = ""
    additional_notes_for_customer: str = ""
    reorder_part: bool = False


def _clean(value):
    return str(value).strip() if value not in (None, "") else ""


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _first_date_from_cells(ws, cells):
    for cell in cells:
        d = _as_date(ws[cell].value)
        if d:
            return d
    return None


def _week_start(any_date):
    # Company timesheet weeks start on Sunday.
    return any_date - timedelta(days=(any_date.weekday() + 1) % 7)


def _find_chunk_date(ws, start_row, end_row, previous_date=None):
    for row in range(start_row, end_row + 1):
        d = _as_date(ws[f"{DATE_COL}{row}"].value)
        if d:
            return d
    if previous_date:
        return previous_date + timedelta(days=1)
    return None


def parse_time_entries(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[TIME_SHEET_NAME]
    parsed = []
    current_date = None

    for start_row, end_row in TIME_ENTRY_CHUNKS:
        current_date = _find_chunk_date(ws, start_row, end_row, current_date)
        if not current_date:
            continue

        # Overnight is a date-group-level value stored on the last row of the group.
        overnight_stay = parse_bool_cell(ws[f"{OVERNIGHT_COL}{end_row}"].value)

        for row in range(start_row, end_row + 1):
            row_order = row - start_row + 1
            job_number = _clean(ws[f"{JOB_COL}{row}"].value)
            work_code = _clean(ws[f"{WORK_CODE_COL}{row}"].value)
            description = _clean(ws[f"{DESCRIPTION_COL}{row}"].value)
            regular = as_decimal(ws[f"{REGULAR_COL}{row}"].value)
            overtime = as_decimal(ws[f"{OVERTIME_COL}{row}"].value)
            doubletime = as_decimal(ws[f"{DOUBLETIME_COL}{row}"].value)

            if not any([job_number, work_code, description, regular, overtime, doubletime]):
                continue

            parsed.append(
                ParsedTimeEntry(
                    work_date=current_date,
                    row_order=row_order,
                    job_number=job_number,
                    work_code=work_code,
                    regular_hours=regular,
                    overtime_hours=overtime,
                    doubletime_hours=doubletime,
                    overnight_stay=overnight_stay,
                    description=description,
                )
            )
    return parsed


def _time_row_to_date_and_order(path):
    """Return a mapping of Time Sheet row -> (work_date, row_order)."""
    wb = load_workbook(path, data_only=True)
    ws = wb[TIME_SHEET_NAME]
    mapping = {}
    current_date = None

    for start_row, end_row in TIME_ENTRY_CHUNKS:
        current_date = _find_chunk_date(ws, start_row, end_row, current_date)
        if not current_date:
            continue
        for row in range(start_row, end_row + 1):
            mapping[row] = (current_date, row - start_row + 1)

    return mapping


def parse_expense_entries(path):
    """Parse the Expense Report sheet.

    Expense Report rows 9-43 are aligned with Time Sheet rows 20-54.
    The workbook's Mileage dollar column is calculated, so the app imports only
    editable values and recalculates mileage from miles * timesheet.mileage_rate.
    """
    wb = load_workbook(path, data_only=True)
    if EXPENSE_SHEET_NAME not in wb.sheetnames:
        return []

    ws = wb[EXPENSE_SHEET_NAME]
    time_row_map = _time_row_to_date_and_order(path)
    parsed = []

    for expense_row in range(EXPENSE_FIRST_ROW, EXPENSE_LAST_ROW + 1):
        time_sheet_row = expense_row + EXPENSE_TIME_ROW_OFFSET
        if time_sheet_row not in time_row_map:
            continue

        work_date, row_order = time_row_map[time_sheet_row]
        item = ParsedExpenseEntry(
            time_sheet_row=time_sheet_row,
            work_date=work_date,
            row_order=row_order,
            miles=as_decimal(ws[f"{EXPENSE_MILES_COL}{expense_row}"].value),
            per_diem_food=as_decimal(ws[f"{EXPENSE_PER_DIEM_FOOD_COL}{expense_row}"].value),
            air_fare=as_decimal(ws[f"{EXPENSE_AIR_FARE_COL}{expense_row}"].value),
            hotel=as_decimal(ws[f"{EXPENSE_HOTEL_COL}{expense_row}"].value),
            tolls_parking=as_decimal(ws[f"{EXPENSE_TOLLS_PARKING_COL}{expense_row}"].value),
            rental_car_fuel=as_decimal(ws[f"{EXPENSE_RENTAL_CAR_FUEL_COL}{expense_row}"].value),
            business_meals=as_decimal(ws[f"{EXPENSE_BUSINESS_MEALS_COL}{expense_row}"].value),
            other_expense=as_decimal(ws[f"{EXPENSE_OTHER_EXPENSE_COL}{expense_row}"].value),
            explanation_of_expenses=_clean(ws[f"{EXPENSE_EXPLANATION_COL}{expense_row}"].value),
        )

        if any([
            item.miles,
            item.per_diem_food,
            item.air_fare,
            item.hotel,
            item.tolls_parking,
            item.rental_car_fuel,
            item.business_meals,
            item.other_expense,
            item.explanation_of_expenses,
        ]):
            parsed.append(item)

    return parsed


def parse_part_entries(path):
    """Parse the Parts Report sheet.

    Parts Report rows 9-43 are aligned with Time Sheet rows 20-54.
    Date and job are formula-driven in the workbook, but the app imports the
    displayed EE Stock/Job # as a snapshot value when present.
    """
    wb = load_workbook(path, data_only=True)
    if PARTS_SHEET_NAME not in wb.sheetnames:
        return []

    ws = wb[PARTS_SHEET_NAME]
    time_row_map = _time_row_to_date_and_order(path)
    parsed = []

    for parts_row in range(PARTS_FIRST_ROW, PARTS_LAST_ROW + 1):
        time_sheet_row = parts_row + PARTS_TIME_ROW_OFFSET
        if time_sheet_row not in time_row_map:
            continue

        work_date, row_order = time_row_map[time_sheet_row]
        item = ParsedPartEntry(
            time_sheet_row=time_sheet_row,
            work_date=work_date,
            row_order=row_order,
            ee_stock_job_number=_clean(ws[f"{PARTS_EE_STOCK_JOB_COL}{parts_row}"].value),
            quantity=as_decimal(ws[f"{PARTS_QUANTITY_COL}{parts_row}"].value),
            part_description_part_number=_clean(ws[f"{PARTS_DESCRIPTION_PART_NUMBER_COL}{parts_row}"].value),
            additional_notes_for_customer=_clean(ws[f"{PARTS_ADDITIONAL_NOTES_COL}{parts_row}"].value),
            reorder_part=parse_bool_cell(ws[f"{PARTS_REORDER_COL}{parts_row}"].value),
        )

        if any([
            item.ee_stock_job_number,
            item.quantity,
            item.part_description_part_number,
            item.additional_notes_for_customer,
            item.reorder_part,
        ]):
            parsed.append(item)

    return parsed


def parse_week_start(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[TIME_SHEET_NAME]
    explicit = _first_date_from_cells(ws, WEEK_START_CELLS)
    if explicit:
        return _week_start(explicit)
    entries = parse_time_entries(path)
    if not entries:
        raise ValueError("Could not determine week start from workbook.")
    return _week_start(entries[0].work_date)



def valid_time_entry_job_qs():
    """Jobs users may select/charge time to during imports and entry forms."""
    return Job.objects.filter(active=True).exclude(description="")


def find_invalid_time_entry_job_numbers(path):
    """Return invalid workbook job numbers grouped with sample entry descriptions.

    A blank job number is valid internal work. A nonblank job number is valid only
    when it exists in the Job table, has a description, and is marked available for
    time entry.
    """
    invalid = {}
    valid_numbers = {
        (value or "").strip().lower()
        for value in valid_time_entry_job_qs().values_list("job_number", flat=True)
    }

    for item in parse_time_entries(path):
        job_number = (item.job_number or "").strip()
        if not job_number:
            continue
        if job_number.lower() in valid_numbers:
            continue

        row = invalid.setdefault(job_number, {"count": 0, "descriptions": []})
        row["count"] += 1
        description = (item.description or "").strip()
        if description and description not in row["descriptions"] and len(row["descriptions"]) < 5:
            row["descriptions"].append(description)

    return invalid


def _resolve_import_job(job_number, job_corrections=None):
    """Resolve a workbook job number to a valid Job or None.

    job_corrections maps original workbook job numbers to either a replacement
    job number or an empty string for internal/no-job work.
    """
    job_number = (job_number or "").strip()
    if not job_number:
        return "", None

    corrections = job_corrections or {}
    if job_number in corrections:
        replacement = (corrections[job_number] or "").strip()
        if not replacement:
            return "", None
        job_number = replacement

    job = valid_time_entry_job_qs().filter(job_number__iexact=job_number).first()
    if job is None:
        raise ValueError(
            f"Job '{job_number}' is not available for time entry. "
            "Import stopped so the job number can be corrected."
        )
    return job.job_number, job

@transaction.atomic
def import_timesheet_upload(upload, job_corrections=None):
    path = upload.uploaded_file.path
    week_start = parse_week_start(path)
    timesheet, _ = Timesheet.objects.get_or_create(
        employee=upload.employee,
        week_start=week_start,
        defaults={"entries_per_day": 5, "template_entries_per_day": 5, "mileage_rate": MileageRate.rate_for_date(week_start)},
    )
    if timesheet.status in {Timesheet.Status.SUBMITTED, Timesheet.Status.APPROVED, Timesheet.Status.INVOICED, Timesheet.Status.VOID}:
        raise ValueError("Submitted, approved, invoiced, or voided timesheets cannot be replaced by upload. Reopen the submitted timesheet first if corrections are needed.")

    # Re-upload replaces the existing week.
    timesheet.entries.all().delete()
    PartEntry.objects.filter(time_entry__timesheet=timesheet).delete()

    entries_by_date_row = {}
    for item in parse_time_entries(path):
        job_number, job = _resolve_import_job(item.job_number, job_corrections)
        work_code = None
        if item.work_code:
            work_code, _ = WorkCode.objects.get_or_create(code=item.work_code, defaults={"description": item.work_code})
        entry = TimeEntry.objects.create(
            timesheet=timesheet,
            work_date=item.work_date,
            row_order=item.row_order,
            job_number=job_number,
            job=job,
            work_code=work_code,
            regular_hours=item.regular_hours,
            overtime_hours=item.overtime_hours,
            doubletime_hours=item.doubletime_hours,
            overnight_stay=item.overnight_stay,
            description=item.description,
        )
        entries_by_date_row[(item.work_date, item.row_order)] = entry

    for item in parse_expense_entries(path):
        entry = entries_by_date_row.get((item.work_date, item.row_order))
        if entry is None:
            # Expenses are one-to-one with TimeEntry. If the spreadsheet contains
            # an expense on a row without a time entry, create a minimal time
            # entry in the matching date/row slot so the expense is not lost.
            entry = TimeEntry.objects.create(
                timesheet=timesheet,
                work_date=item.work_date,
                row_order=item.row_order,
                job_number="",
                regular_hours=Decimal("0"),
                overtime_hours=Decimal("0"),
                doubletime_hours=Decimal("0"),
                description="",
            )
            entries_by_date_row[(item.work_date, item.row_order)] = entry

        Expense.objects.update_or_create(
            time_entry=entry,
            defaults={
                "miles": item.miles,
                "per_diem_food": item.per_diem_food,
                "air_fare": item.air_fare,
                "hotel": item.hotel,
                "tolls_parking": item.tolls_parking,
                "rental_car_fuel": item.rental_car_fuel,
                "business_meals": item.business_meals,
                "other_expense": item.other_expense,
                "explanation_of_expenses": item.explanation_of_expenses,
            },
        )

    for item in parse_part_entries(path):
        entry = entries_by_date_row.get((item.work_date, item.row_order))
        if entry is None:
            # Preserve parts entered on a row without time data.
            part_job_number, part_job = _resolve_import_job(item.ee_stock_job_number, job_corrections)
            entry = TimeEntry.objects.create(
                timesheet=timesheet,
                work_date=item.work_date,
                row_order=item.row_order,
                job_number=part_job_number,
                job=part_job,
                regular_hours=Decimal("0"),
                overtime_hours=Decimal("0"),
                doubletime_hours=Decimal("0"),
                description="",
            )
            entries_by_date_row[(item.work_date, item.row_order)] = entry

        PartEntry.objects.update_or_create(
            time_entry=entry,
            defaults={
                "ee_stock_job_number": item.ee_stock_job_number,
                "quantity": item.quantity,
                "part_description_part_number": item.part_description_part_number,
                "additional_notes_for_customer": item.additional_notes_for_customer,
                "reorder_part": item.reorder_part,
            },
        )


    timesheet.status = Timesheet.Status.DRAFT
    timesheet.submitted_at = None
    timesheet.submitted_by = None
    timesheet.submission_export_format = ""
    timesheet.reopened_at = None
    timesheet.reopened_by = None
    timesheet.reopen_reason = ""
    timesheet.approved_at = None
    timesheet.approved_by = None
    timesheet.invoiced_at = None
    timesheet.invoiced_by = None
    timesheet.deleted_at = None
    timesheet.deleted_by = None
    timesheet.delete_reason = ""
    timesheet.entries_per_day = max(5, timesheet.entries_per_day or 5)
    timesheet.save()

    upload.imported_timesheet = timesheet
    upload.status = "imported"
    upload.message = f"Imported week of {week_start}. Existing records for the week were replaced."
    upload.save(update_fields=["imported_timesheet", "status", "message"])
    return timesheet
