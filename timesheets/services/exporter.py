from pathlib import Path
from decimal import Decimal
from django.conf import settings
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from .workbook_mapping import *

TEMPLATE_PATH = Path(settings.BASE_DIR) / "templates_excel" / "2026_Timesheet.xlsx"

def _export_initials_filename(timesheet, extension):
    employee = timesheet.employee
    user = getattr(employee, "user", employee)

    initials = (
        f"{(user.first_name or '')[:1]}"
        f"{(user.last_name or '')[:1]}"
    ).upper()

    if not initials.strip():
        initials = user.get_username()[:2].upper()

    return f"{timesheet.week_start:%Y%m%d}_{initials}.{extension}"

def _money_amount(value):
    if value is None:
        return Decimal("0")
    return getattr(value, "amount", value) or Decimal("0")



def _excel_number(value):
    amount = _money_amount(value)
    if amount in (None, "", 0, Decimal("0")):
        return ""
    return float(amount)


def _display_number(value):
    amount = _money_amount(value)
    if amount in (None, "", 0, Decimal("0")):
        return ""
    return str(amount)


def _display_money(value):
    amount = _money_amount(value)
    if amount in (None, "", 0, Decimal("0")):
        return ""
    return f"${amount:.2f}"


def _employee_profile(user):
    return getattr(user, "employee_profile", None)


def _write_employee_header(ws, user, timesheet):
    """Populate the new template's employee/office header fields.

    These are database-backed metadata fields and do not affect XLSX import;
    uploaded workbooks still provide the time/expense row data.
    """
    profile = _employee_profile(user)
    full_name = user.get_full_name() or user.get_username()
    first_name = user.first_name or full_name.split(" ")[0]
    last_name = user.last_name or " ".join(full_name.split(" ")[1:])

    office = getattr(profile, "office_location", None) if profile else None

    # New template variable cells.
    ws["M3"] = full_name
    ws["F4"] = office.address_line if office else ""
    ws["F5"] = office.city_state_zip if office else ""
    ws["M4"] = profile.address_line if profile else ""
    ws["M5"] = profile.city_state_zip if profile else ""
    ws["F7"] = timesheet.week_start

    # Backward-compatible cell writes for older templates.
    for cell in FIRST_NAME_CELLS:
        if cell in ws:
            ws[cell] = first_name
            break
    for cell in LAST_NAME_CELLS:
        if cell in ws:
            ws[cell] = last_name
            break



def _has_part_entry(part_entry):
    if not part_entry:
        return False

    quantity = _money_amount(getattr(part_entry, "quantity", 0))

    # Ignore EE Stock/Job # by itself because it is often auto-populated.
    # Only print the Parts Report page when actual part information exists.
    return any([
        quantity,
        getattr(part_entry, "part_description_part_number", ""),
        getattr(part_entry, "additional_notes_for_customer", ""),
        getattr(part_entry, "reorder_part", False),
    ])


def _has_expense(expense):
    if not expense:
        return False
    return any([
        expense.miles,
        _money_amount(expense.per_diem_food),
        _money_amount(expense.air_fare),
        _money_amount(expense.hotel),
        _money_amount(expense.tolls_parking),
        _money_amount(expense.rental_car_fuel),
        _money_amount(expense.business_meals),
        _money_amount(expense.other_expense),
        expense.explanation_of_expenses,
    ])


def build_timesheet_excel(timesheet):
    if not timesheet.can_export_excel:
        raise ValueError("Excel export only works when each date has 5 or fewer time entries. Use PDF Report instead.")
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Missing Excel template: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[TIME_SHEET_NAME]
    expense_ws = wb[EXPENSE_SHEET_NAME] if EXPENSE_SHEET_NAME in wb.sheetnames else None
    parts_ws = wb[PARTS_SHEET_NAME] if PARTS_SHEET_NAME in wb.sheetnames else None
    user = timesheet.employee

    # Populate employee/office header metadata and the official week-start date cell.
    # Column A dates are formulas based on F7 in the new template.
    _write_employee_header(ws, user, timesheet)

    entries_by_date = {}
    for entry in timesheet.entries.select_related("job", "work_code", "expense", "part_entry").order_by("work_date", "row_order"):
        entries_by_date.setdefault(entry.work_date, []).append(entry)

    for chunk_index, (start_row, end_row) in enumerate(TIME_ENTRY_CHUNKS):
        work_date = timesheet.week_start + __import__("datetime").timedelta(days=chunk_index)
        rows = entries_by_date.get(work_date, [])[: timesheet.template_entries_per_day]
        # Do not overwrite the date formulas in column A. They derive from F7.
        overnight_stay = any(entry.overnight_stay for entry in rows)
        ws[f"{OVERNIGHT_COL}{end_row}"] = "=TRUE()" if overnight_stay else "=FALSE()"
        for offset, entry in enumerate(rows):
            row = start_row + offset
            ws[f"{JOB_COL}{row}"] = entry.job_display
            ws[f"{WORK_CODE_COL}{row}"] = entry.work_code.code if entry.work_code else ""
            ws[f"{REGULAR_COL}{row}"] = _excel_number(entry.regular_hours)
            ws[f"{OVERTIME_COL}{row}"] = _excel_number(entry.overtime_hours)
            ws[f"{DOUBLETIME_COL}{row}"] = _excel_number(entry.doubletime_hours)
            ws[f"{DESCRIPTION_COL}{row}"] = entry.description

            # Expense Report rows map one-to-one to Time Sheet rows using the template offset.
            if expense_ws is not None:
                expense_row = row - EXPENSE_TIME_ROW_OFFSET
                if EXPENSE_FIRST_ROW <= expense_row <= EXPENSE_LAST_ROW:
                    expense = getattr(entry, "expense", None)
                    if expense:
                        expense_ws[f"{EXPENSE_MILES_COL}{expense_row}"] = _excel_number(expense.miles)
                        # Mileage amount is calculated, not user editable. Use this timesheet's snapshot rate.
                        expense_ws[f"D{expense_row}"] = f'=IF(C{expense_row}="","",C{expense_row}*{timesheet.mileage_rate})'
                        expense_ws[f"{EXPENSE_PER_DIEM_FOOD_COL}{expense_row}"] = _excel_number(expense.per_diem_food)
                        expense_ws[f"{EXPENSE_AIR_FARE_COL}{expense_row}"] = _excel_number(expense.air_fare)
                        expense_ws[f"{EXPENSE_HOTEL_COL}{expense_row}"] = _excel_number(expense.hotel)
                        expense_ws[f"{EXPENSE_TOLLS_PARKING_COL}{expense_row}"] = _excel_number(expense.tolls_parking)
                        expense_ws[f"{EXPENSE_RENTAL_CAR_FUEL_COL}{expense_row}"] = _excel_number(expense.rental_car_fuel)
                        expense_ws[f"{EXPENSE_BUSINESS_MEALS_COL}{expense_row}"] = _excel_number(expense.business_meals)
                        expense_ws[f"{EXPENSE_OTHER_EXPENSE_COL}{expense_row}"] = _excel_number(expense.other_expense)
                        expense_ws[f"{EXPENSE_EXPLANATION_COL}{expense_row}"] = expense.explanation_of_expenses

            # Parts Report rows map one-to-one to Time Sheet rows using the template offset.
            if parts_ws is not None:
                parts_row = row - PARTS_TIME_ROW_OFFSET
                if PARTS_FIRST_ROW <= parts_row <= PARTS_LAST_ROW:
                    part_entry = getattr(entry, "part_entry", None)
                    if part_entry:
                        # Date and Job # are formula-driven in the workbook.
                        parts_ws[f"{PARTS_QUANTITY_COL}{parts_row}"] = _excel_number(part_entry.quantity)
                        parts_ws[f"{PARTS_DESCRIPTION_PART_NUMBER_COL}{parts_row}"] = part_entry.part_description_part_number
                        parts_ws[f"{PARTS_ADDITIONAL_NOTES_COL}{parts_row}"] = part_entry.additional_notes_for_customer
                        parts_ws[f"{PARTS_REORDER_COL}{parts_row}"] = "=TRUE()" if part_entry.reorder_part else "=FALSE()"

    
    employee = timesheet.employee
    user = getattr(employee, "user", employee)
    initials = (
        f"{(user.first_name or '')[:1]}"
        f"{(user.last_name or '')[:1]}"
    ).upper()

    if not initials.strip():
        initials = user.username[:2].upper()
    # end if

    out_dir = Path(settings.MEDIA_ROOT) / "submitted_timesheets"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / _export_initials_filename(timesheet, "xlsx")

    if out_path.exists():
        out_path.unlink()
    # end if
    wb.save(out_path)
    return out_path



def _receipt_flowables(timesheet, styles):
    """Return ReportLab flowables for uploaded receipt image files.

    PDF receipts are merged into the final PDF after ReportLab builds the main
    report because ReportLab cannot embed PDF pages directly as flowables.
    """
    flowables = []
    image_suffixes = {".jpg", ".jpeg", ".png", ".webp"}

    try:
        receipts = timesheet.receipts.all()
    except Exception:
        return flowables

    for receipt in receipts:
        file_path = getattr(receipt.file, "path", None)
        suffix = Path(receipt.file.name).suffix.lower()

        if suffix not in image_suffixes:
            continue

        if not file_path or not Path(file_path).exists():
            continue

        flowables.append(PageBreak())
        flowables.append(Paragraph("Receipt", styles["Title"]))
        flowables.append(Paragraph(f"File: {receipt.filename()}", styles["Normal"]))

        if receipt.description:
            flowables.append(Paragraph(f"Description: {receipt.description}", styles["Normal"]))

        if receipt.uploaded_at:
            flowables.append(Paragraph(f"Uploaded: {receipt.uploaded_at:%m/%d/%Y %I:%M %p}", styles["Normal"]))

        flowables.append(Spacer(1, 12))

        # Landscape letter with 0.5 inch margins gives 720 x 540 usable points.
        img = Image(file_path)
        max_width = 720
        max_height = 470

        scale = min(max_width / img.imageWidth, max_height / img.imageHeight, 1)
        img.drawWidth = img.imageWidth * scale
        img.drawHeight = img.imageHeight * scale

        flowables.append(img)

    return flowables


def _append_pdf_receipts(report_path, timesheet):
    """Append uploaded PDF receipts to the generated timesheet PDF.

    Image receipts are handled inside ReportLab. PDF receipts are appended here.
    If pypdf is not installed, the main report still exports successfully.
    """
    try:
        from pypdf import PdfReader, PdfWriter
    except Exception:
        return report_path

    try:
        receipts = list(timesheet.receipts.all())
    except Exception:
        return report_path

    pdf_receipts = [
        receipt for receipt in receipts
        if Path(receipt.file.name).suffix.lower() == ".pdf"
        and getattr(receipt.file, "path", None)
        and Path(receipt.file.path).exists()
    ]

    if not pdf_receipts:
        return report_path

    writer = PdfWriter()

    with open(report_path, "rb") as main_file:
        main_reader = PdfReader(main_file)
        for page in main_reader.pages:
            writer.add_page(page)

        for receipt in pdf_receipts:
            try:
                receipt_reader = PdfReader(receipt.file.path)
                for page in receipt_reader.pages:
                    writer.add_page(page)
            except Exception:
                # Skip unreadable PDF attachments rather than failing the report.
                continue

        tmp_path = Path(str(report_path) + ".tmp")
        with open(tmp_path, "wb") as out_file:
            writer.write(out_file)

    tmp_path.replace(report_path)
    return report_path


def build_timesheet_pdf(timesheet):

    employee = timesheet.employee
    user = getattr(employee, "user", employee)
    initials = (
        f"{(user.first_name or '')[:1]}"
        f"{(user.last_name or '')[:1]}"
    ).upper()

    if not initials.strip():
        initials = user.username[:2].upper()
    # end if

    out_dir = Path(settings.MEDIA_ROOT) / "submitted_timesheets"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / _export_initials_filename(timesheet, "pdf")
    if out_path.exists():
        out_path.unlink()
    # end if

    doc = SimpleDocTemplate(str(out_path), pagesize=landscape(letter), leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=7, leading=8)
    small_wrap = ParagraphStyle("small_wrap", parent=small, wordWrap="CJK")
    story = []
    employee_name = timesheet.employee.get_full_name() or timesheet.employee.get_username()

    story.append(Paragraph("Timesheet Report", styles["Title"]))
    story.append(Paragraph(f"Employee: {employee_name}", styles["Normal"]))
    story.append(Paragraph(f"Week Start: {timesheet.week_start:%m/%d/%Y}", styles["Normal"]))
    story.append(Paragraph(f"Mileage Rate: ${timesheet.mileage_rate} per mile", styles["Normal"]))
    if timesheet.submitted_at:
        story.append(Paragraph(f"Submitted: {timesheet.submitted_at:%m/%d/%Y %I:%M %p}", styles["Normal"]))
    story.append(Spacer(1, 12))

    data = [["Date", "Job", "Code", "Reg", "OT", "DT", "Hotel", "Description"]]
    total_reg = total_ot = total_dt = 0
    for entry in timesheet.entries.select_related("job", "work_code").order_by("work_date", "row_order"):
        total_reg += entry.regular_hours
        total_ot += entry.overtime_hours
        total_dt += entry.doubletime_hours
        data.append([
            entry.work_date.strftime("%m/%d/%Y"),
            Paragraph(entry.job_display or "", small),
            Paragraph(entry.work_code.code if entry.work_code else "", small),
            str(entry.regular_hours),
            str(entry.overtime_hours),
            str(entry.doubletime_hours),
            "Yes" if entry.overnight_stay else "No",
            Paragraph(entry.description or "", small_wrap),
        ])

    timetable_colWidths=[
        55,   # Date
        70,   # Job
        45,   # Code
        35,   # Reg
        35,   # OT
        35,   # DT
        40,   # Hotel
        405,  # Description
    ]

    # sets up the table column widths
    expensetable_colWidths=[
        50,   # Date
        58,   # Job
        38,   # Miles
        48,   # Mileage
        48,   # Per Diem
        40,   # Air
        40,   # Hotel
        58,   # Tolls/Parking
        58,   # Rental/Fuel
        45,   # Meals
        45,   # Other
        192,  # Explanation
    ]

    table = Table(data, repeatRows=1, colWidths=timetable_colWidths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Totals — Regular: {total_reg} | OT: {total_ot} | DT: {total_dt}", styles["Heading3"]))

    expenses = []
    for entry in timesheet.entries.select_related("expense").order_by("work_date", "row_order"):
        expense = getattr(entry, "expense", None)
        if not _has_expense(expense):
            continue
        expenses.append([
            entry.work_date.strftime("%m/%d/%Y"),
            Paragraph(entry.job_display or "", small),
            str(expense.miles),
            f"${expense.mileage.amount:.2f}",
            f"${expense.per_diem_food.amount:.2f}",
            f"${expense.air_fare.amount:.2f}",
            f"${expense.hotel.amount:.2f}",
            f"${expense.tolls_parking.amount:.2f}",
            f"${expense.rental_car_fuel.amount:.2f}",
            f"${expense.business_meals.amount:.2f}",
            f"${expense.other_expense.amount:.2f}",
            Paragraph(expense.explanation_of_expenses or "", small_wrap),
        ])

    if expenses:
        story.append(PageBreak())
        story.append(Paragraph("Expense Report", styles["Title"]))
        story.append(Paragraph(f"Employee: {employee_name}", styles["Normal"]))
        story.append(Paragraph(f"Week Start: {timesheet.week_start:%m/%d/%Y}", styles["Normal"]))
        story.append(Spacer(1, 12))
        expense_data = [["Date", "Job", "Miles", "Mileage", "Per Diem", "Air", "Hotel", "Tolls/Parking", "Rental/Fuel", "Meals", "Other", "Explanation"]] + expenses
        expense_table = Table(expense_data, repeatRows=1, colWidths=expensetable_colWidths)
        expense_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(expense_table)

    parts = []
    for entry in timesheet.entries.select_related("part_entry").order_by("work_date", "row_order"):
        part_entry = getattr(entry, "part_entry", None)
        if not part_entry:
            continue

        quantity = _money_amount(getattr(part_entry, "quantity", 0))
        part_description = (getattr(part_entry, "part_description_part_number", "") or "").strip()
        customer_notes = (getattr(part_entry, "additional_notes_for_customer", "") or "").strip()
        reorder_part = bool(getattr(part_entry, "reorder_part", False))

        # Do NOT count EE Stock/Job # as parts data. It can be copied from the
        # time entry and should not cause a blank Parts Report page to print.
        has_actual_part_data = any([
            quantity,
            part_description,
            customer_notes,
            reorder_part,
        ])

        if not has_actual_part_data:
            continue

        parts.append([
            entry.work_date.strftime("%m/%d/%Y"),
            Paragraph(part_entry.job_display or entry.job_display or "", small),
            _display_number(quantity),
            Paragraph(part_description, small_wrap),
            Paragraph(customer_notes, small_wrap),
            "Yes" if reorder_part else "No",
        ])

    # Only print the Parts Report page when at least one non-blank part row exists.
    if parts:
        story.append(PageBreak())
        story.append(Paragraph("Parts Report", styles["Title"]))
        story.append(Paragraph(f"Employee: {employee_name}", styles["Normal"]))
        story.append(Paragraph(f"Week Start: {timesheet.week_start:%m/%d/%Y}", styles["Normal"]))
        story.append(Spacer(1, 12))
        parts_data = [["Date", "EE Stock/Job #", "Quantity", "Part Description/Part Number", "Additional Notes For Customer", "Re-Order"]] + parts
        parts_table = Table(parts_data, repeatRows=1, colWidths=[55, 85, 55, 210, 250, 65])
        parts_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(parts_table)


    story.extend(_receipt_flowables(timesheet, styles))

    doc.build(story)
    _append_pdf_receipts(out_path, timesheet)
    return out_path


def build_submission_attachment(timesheet, export_format):
    if export_format == timesheet.ExportFormat.EXCEL:
        return build_timesheet_excel(timesheet)
    return build_timesheet_pdf(timesheet)
