from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import re
from typing import Any

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from timesheets.models import Customer, Job

HEADER_ALIASES = {
    "job_number": {"quote/job #", "job #", "job number", "quote #/job #"},
    "year": {"year"},
    "job_type": {"job type", "type"},
    "cfr_job_number": {"cfr job#", "cfr job #", "cfr job number"},
    "customer": {"customer"},
    "job_status": {"job status", "status"},
    "invoice_status": {"invoice status"},
    "work_type": {"work type"},
    "location": {"location"},
    "customer_contact": {"customer contact", "contact"},
    "customer_po": {"po#", "po #", "po number", "customer po"},
    "description": {"description"},
    "lead": {"lead"},
    "quote_date": {"quote date"},
    "accepted_date": {"accepted date"},
    "quote_number": {"quote #", "quote number"},
    "comments": {"comments", "notes"},
}

for i in range(1, 11):
    HEADER_ALIASES[f"engineer_{i:02d}"] = {f"engineer{i:02d}", f"engineer {i:02d}", f"engineer{i}", f"engineer {i}"}


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date(value: Any):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _text(value)
    if not raw or raw.upper() == "N/A":
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def _int(value: Any):
    raw = _text(value)
    if not raw:
        return None
    try:
        return int(float(raw))
    except ValueError:
        return None


def _user_map():
    User = get_user_model()
    mapping = {}
    for user in User.objects.all():
        full = _norm(user.get_full_name())
        if full:
            mapping[full] = user
        username = _norm(user.username)
        if username:
            mapping[username] = user
    return mapping


def find_header_row(ws):
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25)):
        values = [_norm(cell.value) for cell in row]
        if any(v in HEADER_ALIASES["job_number"] for v in values):
            return row[0].row
    return 1


def build_header_map(ws, header_row):
    aliases = {alias: key for key, values in HEADER_ALIASES.items() for alias in values}
    header_map = {}
    for cell in ws[header_row]:
        key = aliases.get(_norm(cell.value))
        if key:
            header_map[key] = cell.column
    return header_map


def is_year_separator(row_values, job_number):
    populated = [v for v in row_values if _text(v)]
    raw = _text(job_number)
    return bool(re.fullmatch(r"20\d{2}", raw)) and len(populated) <= 2


@dataclass
class JobImportResult:
    added: int = 0
    updated: int = 0
    unchanged: int = 0
    ignored_blank: int = 0
    ignored_year_rows: int = 0
    ignored_invalid: int = 0
    unknown_leads: set[str] = field(default_factory=set)
    unknown_engineers: set[str] = field(default_factory=set)
    errors: list[str] = field(default_factory=list)

    @property
    def total_changed(self):
        return self.added + self.updated


def preview_job_import(path):
    return import_job_list(path, apply=False)


@transaction.atomic
def apply_job_import(path, *, user=None, source_name=""):
    return import_job_list(path, apply=True, user=user, source_name=source_name)


def import_job_list(path, *, apply=False, user=None, source_name=""):
    result = JobImportResult()
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Jobs - Quotes"] if "Jobs - Quotes" in wb.sheetnames else wb.active
    header_row = find_header_row(ws)
    header_map = build_header_map(ws, header_row)

    if "job_number" not in header_map:
        result.errors.append("Could not find a Job Number column. Expected a header like 'Quote/Job #' or 'Job #'.")
        return result

    users = _user_map()
    now = timezone.now()

    for row in ws.iter_rows(min_row=header_row + 1):
        row_values = [cell.value for cell in row]
        if not any(_text(v) for v in row_values):
            result.ignored_blank += 1
            continue

        def value(key):
            col = header_map.get(key)
            if not col:
                return None
            return row[col - 1].value if col - 1 < len(row) else None

        job_number = _text(value("job_number"))
        if is_year_separator(row_values, job_number):
            result.ignored_year_rows += 1
            continue
        if not job_number or not re.search(r"\d", job_number):
            result.ignored_invalid += 1
            continue

        customer_name = _text(value("customer"))
        lead_name = _text(value("lead"))
        lead_user = users.get(_norm(lead_name)) if lead_name else None
        if lead_name and not lead_user:
            result.unknown_leads.add(lead_name)

        engineer_users = []
        for i in range(1, 11):
            eng_name = _text(value(f"engineer_{i:02d}"))
            if not eng_name:
                continue
            eng_user = users.get(_norm(eng_name))
            if eng_user:
                engineer_users.append(eng_user)
            else:
                result.unknown_engineers.add(eng_name)

        fields = {
            "description": _text(value("description")),
            "year": _int(value("year")),
            "job_type": _text(value("job_type")),
            "cfr_job_number": _text(value("cfr_job_number")),
            "job_status": _text(value("job_status")) or Job.STATUS_UNKNOWN,
            "invoice_status": _text(value("invoice_status")),
            "work_type": _text(value("work_type")),
            "location": _text(value("location")),
            "customer_contact": _text(value("customer_contact")),
            "customer_po": _text(value("customer_po")),
            "lead": lead_name,
            "lead_user": lead_user,
            "quote_date": _date(value("quote_date")),
            "accepted_date": _date(value("accepted_date")),
            "quote_number": _text(value("quote_number")),
            "comments": _text(value("comments")),
            "import_source": source_name,
            "last_imported_at": now,
            "active": True,
        }
        for i in range(1, 11):
            fields[f"engineer_{i:02d}"] = _text(value(f"engineer_{i:02d}"))

        job = Job.objects.filter(job_number__iexact=job_number).first()
        if job is None:
            result.added += 1
            if apply:
                customer = None
                if customer_name:
                    customer, _ = Customer.objects.get_or_create(name=customer_name)
                job = Job(job_number=job_number, customer=customer, **fields)
                job.save()
                if engineer_users:
                    job.engineer_users.set(engineer_users)
            continue

        changed = False
        if customer_name:
            customer, _ = Customer.objects.get_or_create(name=customer_name) if apply else (None, False)
            current_customer_name = job.customer.name if job.customer else ""
            if current_customer_name != customer_name:
                changed = True
                if apply:
                    job.customer = customer

        for key, new_value in fields.items():
            if new_value in (None, "") and key not in {"invoice_status", "comments", "lead", "lead_user"}:
                continue
            if getattr(job, key, None) != new_value:
                changed = True
                if apply:
                    setattr(job, key, new_value)

        existing_engineer_ids = set(job.engineer_users.values_list("id", flat=True)) if job.pk else set()
        new_engineer_ids = {u.id for u in engineer_users}
        if new_engineer_ids and existing_engineer_ids != new_engineer_ids:
            changed = True

        if changed:
            result.updated += 1
            if apply:
                job.save()
                if new_engineer_ids:
                    job.engineer_users.set(engineer_users)
        else:
            result.unchanged += 1

    return result
