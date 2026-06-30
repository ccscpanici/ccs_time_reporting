from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import re

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from timesheets.models import Customer, Job, infer_job_year_and_month


YEAR_SEPARATOR_RE = re.compile(r"^20\d{2}$")
SHEET_NAME = "Jobs - Quotes"
HEADER_ROW = 2

FIELD_MAP = {
    "Quote/Job #": "job_number",
    "Year": "year",
    "Job Type": "job_type",
    "CFR Job#": "cfr_job_number",
    "Customer": "customer_name",
    "Job Status": "job_status",
    "Invoice Status": "invoice_status",
    "Work Type": "work_type",
    "Location": "location",
    "Customer Contact": "customer_contact",
    "PO#": "customer_po",
    "Description": "description",
    "Lead": "lead",
    "Quote Date": "quote_date",
    "Accepted Date": "accepted_date",
    "Quote #": "quote_number",
    "Comments": "comments",
    "Engineer01": "engineer_01",
    "Engineer02": "engineer_02",
    "Engineer03": "engineer_03",
    "Engineer04": "engineer_04",
    "Engineer05": "engineer_05",
    "Engineer06": "engineer_06",
    "Engineer07": "engineer_07",
    "Engineer08": "engineer_08",
    "Engineer09": "engineer_09",
    "Engineer10": "engineer_10",
}

TEXT_FIELDS = {
    "job_type",
    "cfr_job_number",
    "job_status",
    "invoice_status",
    "work_type",
    "location",
    "customer_contact",
    "customer_po",
    "description",
    "lead",
    "quote_number",
    "comments",
    "engineer_01",
    "engineer_02",
    "engineer_03",
    "engineer_04",
    "engineer_05",
    "engineer_06",
    "engineer_07",
    "engineer_08",
    "engineer_09",
    "engineer_10",
}


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_job_number(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_year(value):
    if value in (None, ""):
        return None
    try:
        year = int(Decimal(str(value)).quantize(Decimal("1")))
        if 0 <= year < 100:
            return 2000 + year
        return year
    except (InvalidOperation, ValueError):
        return None


def clean_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                pass
    return None


def is_year_separator_row(job_number, values, customer_name):
    """Return True for orange year-divider rows such as a lone 2026 row."""
    if not YEAR_SEPARATOR_RE.match(job_number):
        return False
    meaningful_fields = [
        customer_name,
        values.get("description"),
        values.get("job_status"),
        values.get("invoice_status"),
        values.get("work_type"),
        values.get("location"),
        values.get("customer_po"),
        values.get("lead"),
        values.get("quote_number"),
        values.get("comments"),
    ]
    meaningful_fields.extend(values.get(f"engineer_{idx:02d}") for idx in range(1, 11))
    return not any(clean_text(value) for value in meaningful_fields)


def user_full_name_key(user):
    return f"{user.first_name} {user.last_name}".strip().casefold()


def resolve_user_by_full_name(name, users_by_name):
    raw = clean_text(name)
    if not raw:
        return None
    return users_by_name.get(raw.casefold())


class Command(BaseCommand):
    help = "Import CCS jobs from the company Work Order and Job List workbook."

    def add_arguments(self, parser):
        parser.add_argument("workbook", help="Path to CCS Work Order and Job List .xlsx file")
        parser.add_argument("--sheet", default=SHEET_NAME, help=f"Worksheet name. Default: {SHEET_NAME}")
        parser.add_argument("--dry-run", action="store_true", help="Parse and report counts without saving changes")

    def handle(self, *args, **options):
        workbook_path = options["workbook"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]

        try:
            wb = load_workbook(workbook_path, read_only=False, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"Workbook not found: {workbook_path}") from exc

        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Worksheet '{sheet_name}' was not found. Available sheets: {', '.join(wb.sheetnames)}")

        ws = wb[sheet_name]
        headers = [clean_text(cell.value) for cell in ws[HEADER_ROW]]
        column_map = {}
        for idx, header in enumerate(headers, start=1):
            if header in FIELD_MAP:
                column_map[FIELD_MAP[header]] = idx

        if "job_number" not in column_map:
            raise CommandError("Could not find required 'Quote/Job #' column.")

        created = 0
        updated = 0
        skipped = 0
        unknown_status = 0
        now = timezone.now()
        User = get_user_model()
        users_by_name = {
            user_full_name_key(user): user
            for user in User.objects.filter(is_active=True)
            if user_full_name_key(user)
        }

        @transaction.atomic
        def import_rows():
            nonlocal created, updated, skipped, unknown_status
            for row_number in range(HEADER_ROW + 1, ws.max_row + 1):
                raw_job_number = ws.cell(row_number, column_map["job_number"]).value
                job_number = clean_job_number(raw_job_number)
                if not job_number:
                    skipped += 1
                    continue

                values = {}
                customer_name = ""
                for field, col_idx in column_map.items():
                    value = ws.cell(row_number, col_idx).value
                    if field == "job_number":
                        continue
                    if field == "customer_name":
                        customer_name = clean_text(value)
                    elif field == "year":
                        values[field] = clean_year(value)
                    elif field in {"quote_date", "accepted_date"}:
                        values[field] = clean_date(value)
                    elif field in TEXT_FIELDS:
                        values[field] = clean_text(value)

                if is_year_separator_row(job_number, values, customer_name):
                    skipped += 1
                    continue

                values["lead_user"] = resolve_user_by_full_name(values.get("lead"), users_by_name)
                for idx in range(1, 11):
                    field_name = f"engineer_{idx:02d}"
                    values[f"{field_name}_user"] = resolve_user_by_full_name(values.get(field_name), users_by_name)

                if values.get("year") is None:
                    inferred_year, inferred_month = infer_job_year_and_month(job_number)
                    values["year"] = inferred_year
                    if inferred_month is not None:
                        values["job_month"] = inferred_month
                elif "job_month" not in values:
                    _inferred_year, inferred_month = infer_job_year_and_month(job_number)
                    if inferred_month is not None:
                        values["job_month"] = inferred_month

                if not values.get("job_status"):
                    values["job_status"] = Job.STATUS_UNKNOWN
                    unknown_status += 1

                customer = None
                if customer_name:
                    customer, _ = Customer.objects.get_or_create(name=customer_name)
                values["customer"] = customer
                values["active"] = True
                values["import_source"] = sheet_name
                values["last_imported_at"] = now

                job, was_created = Job.objects.update_or_create(
                    job_number=job_number,
                    defaults=values,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        if dry_run:
            with transaction.atomic():
                import_rows()
                transaction.set_rollback(True)
        else:
            import_rows()

        mode = "DRY RUN - " if dry_run else ""
        self.stdout.write(self.style.SUCCESS(
            f"{mode}Imported jobs from '{sheet_name}': created={created}, updated={updated}, skipped={skipped}, blank_status_as_unknown={unknown_status}"
        ))
